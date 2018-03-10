
# export data sheets from xlsx to csv
# Found this code on Github:
# https://gist.github.com/julianthome/2d8546e7bed869079ab0f409ae0faa87
# By julianthome

from openpyxl import load_workbook
import csv
from os import sys

def get_all_sheets(excel_file):
    sheets = []
    workbook = load_workbook(excel_file,read_only=True,data_only=True)
    all_worksheets = workbook.get_sheet_names()
    for worksheet_name in all_worksheets:
        sheets.append(worksheet_name)
    return sheets

def csv_from_excel(excel_file, sheets):
    workbook = load_workbook(excel_file,data_only=True)
    csv_path = ""
    for worksheet_name in sheets:
        try:
            worksheet = workbook.get_sheet_by_name(worksheet_name)
        except KeyError:
            print("Could not find " + worksheet_name)
            sys.exit(1)
        csv_path = ''.join([worksheet_name,'.csv'])
        your_csv_file = open(csv_path, 'w')
        wr = csv.writer(your_csv_file)
        for row in worksheet.iter_rows():
            lrow = []
            for cell in row:
                lrow.append(cell.value)
            wr.writerow(lrow)
        your_csv_file.close()
    return csv_path

def converter_XLSX_to_CSV(xlsxFile):
    sheets = []
    sheets = get_all_sheets(xlsxFile)
    assert(sheets != None and len(sheets) > 0)
    return csv_from_excel(xlsxFile, sheets)
